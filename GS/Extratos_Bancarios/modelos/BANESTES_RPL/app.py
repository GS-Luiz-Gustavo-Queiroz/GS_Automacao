import pandas as pd
import unicodedata
import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font

def normalizar_texto(texto):
    if not isinstance(texto, str):
        return ""
    texto = unicodedata.normalize('NFKD', texto).encode('ASCII', 'ignore').decode('ASCII')
    return ''.join(c for c in texto if c.isalnum() or c.isspace()).strip().lower()

def criar_nome_arquivo_saida(arquivo_original, nome_planilha):
    base, _ = os.path.splitext(arquivo_original)
    contador = 1
    while True:
        novo_nome = f"{base}_formatado_{nome_planilha}_{contador}.xlsx"
        if not os.path.exists(novo_nome):
            return novo_nome
        contador += 1

def formatar_contabil(valor):
    if pd.isna(valor):
        return ""
    try:
        valor = float(valor)
        return f"{valor:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
    except:
        return str(valor)

def converter_valor(valor):
    try:
        if isinstance(valor, str):
            return float(valor.replace('.', '').replace(',', '.'))
        return float(valor)
    except (ValueError, TypeError):
        return 0.0

def extrair_dados(path):
    try:
        wb = load_workbook(path, data_only=True)
        for sheet_name in wb.sheetnames:
            print(f"\nProcessando planilha: {sheet_name}")
            ws = wb[sheet_name]

            dados = []
            for row in ws.iter_rows(values_only=False):
                if all(cell.value is None for cell in row):
                    continue

                # Verifica se a primeira célula está em negrito
                if row[0].font and row[0].font.bold:
                    valores = [cell.value for cell in row]
                    dados.append(valores)

            if not dados:
                print("Nenhuma linha em negrito encontrada.")
                continue

            df = pd.DataFrame(dados)
            processar_dataframe(df, path, sheet_name)

    except Exception as e:
        print(f"Erro ao processar o arquivo: {e}")

def processar_dataframe(df, arquivo, nome_planilha):
    linha_cabecalho = None
    for idx, linha in df.iterrows():
        linha_normalizada = [normalizar_texto(str(cell)) for cell in linha.values]
        if any('data' in col for col in linha_normalizada) and any('valor' in col for col in linha_normalizada):
            linha_cabecalho = idx
            break

    if linha_cabecalho is None:
        print("Cabeçalhos de data/valor não encontrados.")
        return

    df.columns = df.iloc[linha_cabecalho]
    df = df.drop(index=range(0, linha_cabecalho + 1))
    df = df.dropna(how='all')

    col_data = next((col for col in df.columns if 'data' in normalizar_texto(str(col))), None)
    col_valor = next((col for col in df.columns if 'valor' in normalizar_texto(str(col))), None)

    if not col_data or not col_valor:
        print("Colunas de Data ou Valor não encontradas.")
        return

    df = df[[col_data, col_valor]].rename(columns={col_data: 'Data', col_valor: 'Saldo'})
    df['Data'] = pd.to_datetime(df['Data'], dayfirst=True, errors='coerce')
    df['Saldo'] = df['Saldo'].apply(converter_valor)
    df = df.dropna(subset=['Data'])

    # Ordena por data e extrai a última ocorrência de cada dia
    df = df.sort_values('Data')
    df['Data_Dia'] = df['Data'].dt.date
    df_filtrado = df.groupby('Data_Dia', as_index=False).last()
    df_filtrado['Data'] = pd.to_datetime(df_filtrado['Data_Dia']).dt.strftime('%d/%m/%Y')
    df_filtrado = df_filtrado.drop(columns=['Data_Dia'])
    df_filtrado['Saldo'] = df_filtrado['Saldo'].apply(formatar_contabil)

    # Criar e salvar planilha formatada
    nome_saida = criar_nome_arquivo_saida(arquivo, nome_planilha)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Saldo por Dia'

    for r in dataframe_to_rows(df_filtrado, index=False, header=True):
        ws.append(r)

    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = '#.##0,00_-'
            cell.font = Font(bold=True)

    wb.save(nome_saida)
    print(f"Arquivo salvo com sucesso: {nome_saida}")

def BANESTES_RPL(path: str):
    if os.path.isfile(path) and path.lower().endswith('.xlsx'):
        print(f"\nArquivo recebido: {path}")
        extrair_dados(path)
    else:
        print(f"Arquivo inválido ou não encontrado (apenas .xlsx são aceitos): {path}")
